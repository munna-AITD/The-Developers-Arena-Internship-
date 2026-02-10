import os
from openpyxl import Workbook, load_workbook
from book import Book

BASE_DIR = os.path.dirname(__file__)
DATA_FILE = os.path.join(BASE_DIR, "data", "books.xlsx")


def initialize_excel():
    if not os.path.exists(DATA_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Books"
        ws.append(["Book ID", "Title", "Author", "Status"])
        wb.save(DATA_FILE)


def read_all_books():
    initialize_excel()
    wb = load_workbook(DATA_FILE)
    ws = wb.active

    books = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        book_id, title, author, status = row
        books.append(Book(book_id, title, author, status))
    return books


def write_all_books(books):
    wb = Workbook()
    ws = wb.active
    ws.title = "Books"
    ws.append(["Book ID", "Title", "Author", "Status"])

    for b in books:
        ws.append([b.book_id, b.title, b.author, b.status])

    wb.save(DATA_FILE)


def add_book(book):
    initialize_excel()
    wb = load_workbook(DATA_FILE)
    ws = wb.active
    ws.append([book.book_id, book.title, book.author, book.status])
    wb.save(DATA_FILE)
